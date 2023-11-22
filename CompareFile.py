import pandas as pd
from openpyxl.styles import PatternFill
from datetime import datetime

# Get the current date and time
now = datetime.now()

# Format the date and time
date_time = now.strftime("%m-%d-%Y")
# Load the two Excel files (versions)
# file1 : old file - teams site
# file2 : new file - sync files

file1 = pd.ExcelFile(fr"C:\Users\eg724520\Valmont Industries, Inc\Global IT AppSOD & DP - Documents\Team Tools\PyhtonCode - SAP Workstream DDF Data Refresh\Level 3 Tracking Data\Output\sync_results_Level3TrackingData_11-17-2023.xlsx")

# file2 = pd.ExcelFile(
#     r"C:\Users\eg724520\OneDrive - Valmont Industries, Inc\Documents\DDFCompare\sync_results_processStepDesc_DDF.xlsx")
# file2 = pd.ExcelFile(r"C:\Users\eg724520\OneDrive - Valmont Industries, Inc\Documents\DDFCompare\ProcessStepDesc_1680093583905_Excel.xlsx (4).xlsx")

file2= pd.ExcelFile(fr"C:\Users\eg724520\Valmont Industries, Inc\Global IT AppSOD & DP - Documents\Team Tools\PyhtonCode - SAP Workstream DDF Data Refresh\Level 3 Tracking Data\Output\sync_results_Level3TrackingData_{date_time}.xlsx")


# file1 = pd.ExcelFile(r"C:\Users\eg724520\OneDrive - Valmont Industries, Inc\Documents\DDFCompare\CAP DDF_Level3TrackingData_8-10-23.xlsx")
# file2 = pd.ExcelFile(r"C:\Users\eg724520\OneDrive - Valmont Industries, Inc\Documents\DDFCompare\sync_results_level3trackingdata.xlsx")
# Get the list of sheet names in both files
sheet_names = file1.sheet_names
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
highlight_fill_red = PatternFill(start_color="FF0101", end_color="FFFF00", fill_type="solid")
highlight_fill_green = PatternFill(start_color="92d050", end_color="92d050", fill_type="solid")
key = {'L3_Process_ID', 'Busines Process L3', 'Level 3 Business Process Process Id', 'T_Code', 'Name', 'User ID'}
# key = {'T_Code', 'App ID', 'Step'}
# key = {'Security Group', 'Parent Menu ID', 'Option ID', 'Security Code'}
# change the output path
new_sync_data = {}
with pd.ExcelWriter(
        fr"C:\Users\eg724520\Valmont Industries, Inc\Global IT AppSOD & DP - Documents\Team Tools\PyhtonCode - SAP Workstream DDF Data Refresh\Level 3 Tracking Data\Output\Compare_results_Level3TrackingData_{date_time}.xlsx",
        engine='openpyxl') as writer:
    for sheet_name in sheet_names:
        key_columns = []
        try:
            # Load the sheets as DataFrames
            sheet1 = file1.parse(sheet_name)
            sheet2 = file2.parse(sheet_name)
            for col in key:
                if col in sheet2.columns:
                    key_columns.append(col)

            for col in key:
                if col in sheet2.columns:
                    common_column = col
            outer_join = pd.merge(sheet1, sheet2, indicator=True, how='outer')
            if sheet_name != 'MasterList' and 'Changes' in sheet2.columns and 'Changes' in sheet1.columns:

                outer_join.loc[:,'Changes'] = f'Added {date_time}'
            # Perform left join
            left_join_result = outer_join[outer_join['_merge'] == 'left_only']
            right_join_result = outer_join[outer_join['_merge'] == 'right_only']
            # left_join_result.loc[:,'Changes'] = f'Added {date_time}'
            # right_join_result.loc[:,'Changes'] = f'Added {date_time}'

            # Save the styled left join result to the Excel file
            left_join_result.to_excel(writer, sheet_name=f'O_{sheet_name}', index=False)
            # Get the sheet using openpyxl
            styled_sheet = writer.sheets[f'O_{sheet_name}']

            # Apply cell highlighting based on differences
            # Assuming sheet2 is a DataFrame
            key_columns_mapping = {}
            column_index_mapping = {col: index + 1 for index, col in enumerate(sheet2.columns)}
            for x in column_index_mapping.keys():
                for y in key_columns:
                    if y == x:
                        key_columns_mapping[y] = column_index_mapping[x]

            # Create a dictionary from right_join_result
            right_join_dict = {tuple(row[col] for col in key_columns): row for _, row in right_join_result.iterrows()}

            # Convert key_columns_mapping to a set for faster lookups
            key_columns_set = set(sheet2[[y for y in key_columns_mapping.keys()]].values.flatten())

            # Now iterate over styled_sheet
            for row in styled_sheet.iter_rows(min_row=2, max_row=styled_sheet.max_row, min_col=1,
                                              max_col=styled_sheet.max_column):
                styled_row_values = tuple(
                    row[column_index_mapping[col] - 1].value for col in key_columns if col in column_index_mapping)

                # Look up the corresponding row in right_join_dict
                right_row = right_join_dict.get(styled_row_values)

                # If there is a corresponding row, compare the values
                if right_row is not None:
                    change = False
                    x = 0
                    for cell, right_value in zip(row, right_row):
                        col_index = cell.column - 1  # Adjust for 0-based indexing
                        if not pd.isnull(
                                cell.value) and cell.value != right_value and cell.column != styled_sheet.max_column - 1 and cell.value!= '' and right_value != '':
                            cell.fill = highlight_fill
                            change = True
                            if sheet_name == "MasterList" :
                                x = cell.column

                        if cell.column == x + 1 and change and cell.column % 2 == 0 and cell.column >= 4:
                            cell.value = f"Updated {date_time}"
                        if cell.column == styled_sheet.max_column - 1 and change and sheet_name != "MasterList" and 'Changes' in sheet2.columns and 'Changes' in sheet1.columns:
                            cell.value = f'Updated {date_time}'

                flag = False
                for cell in row:
                    if cell.column in key_columns_mapping.values():
                        if cell.value not in key_columns_set:
                            flag = True
                            cell.fill = highlight_fill_red  # highlight the key that does not exist in the other file
                    if cell.column == styled_sheet.max_column - 1 and flag and 'Changes' in sheet2.columns and 'Changes' in sheet1.columns:
                        cell.value = f'Removed {date_time}'
                    if cell.column == styled_sheet.max_column - 1 and 'Added' in cell.value and 'Changes' in sheet2.columns and 'Changes' in sheet1.columns:
                        cell.value = f'Removed {date_time}'
            # Save the styled right join result to the Excel file
            right_join_result.to_excel(writer, sheet_name=f'N_{sheet_name}', index=False)
            styled_sheet = writer.sheets[f'N_{sheet_name}']
            print(len(sheet2[[y for y in key_columns_mapping.keys()]].values))



            # Apply cell highlighting based on differences
            # Create a dictionary from left_join_result
            left_join_dict = {tuple(row[col] for col in key_columns): row for _, row in left_join_result.iterrows()}

            # Convert key_columns_mapping to a set for faster lookups
            key_columns_set = set(sheet1[[y for y in key_columns_mapping.keys()]].values.flatten())

            # Now iterate over styled_sheet
            for row in styled_sheet.iter_rows(min_row=2, max_row=styled_sheet.max_row, min_col=1,
                                              max_col=styled_sheet.max_column):
                styled_row_values = tuple(
                    row[column_index_mapping[col] - 1].value for col in key_columns if col in column_index_mapping)

                # Look up the corresponding row in left_join_dict
                left_row = left_join_dict.get(styled_row_values)

                # If there is a corresponding row, compare the values
                if left_row is not None:
                    change = False
                    x = 0
                    for cell, left_value in zip(row, left_row):
                        col_index = cell.column - 1  # Adjust for 0-based indexing
                        if not pd.isnull(
                                cell.value) and cell.value != left_value and cell.column != styled_sheet.max_column - 1 and cell.value!= '' and left_value != '':
                            change = True
                            cell.fill = highlight_fill
                            if sheet_name == "MasterList":
                                x = cell.column
                        if cell.column == x + 1 and change and cell.column % 2 == 0 and cell.column >= 4:

                            cell.value = f"Updated {date_time}"

                        if cell.column == styled_sheet.max_column - 1 and change and sheet_name != "MasterList" and 'Changes' in sheet2.columns and 'Changes' in sheet1.columns:
                            cell.value = f'Updated {date_time}'
                flag = False
                for cell in row:
                    if cell.column in key_columns_mapping.values():
                        if cell.value not in key_columns_set:
                            flag = True
                            cell.fill = highlight_fill_green
                    if cell.column == styled_sheet.max_column - 1 and flag and 'Changes' in sheet2.columns and 'Changes' in sheet1.columns:
                        cell.value = f'Added {date_time}'

            data = []

            # Iterate through the rows of the styled sheet
            for row in styled_sheet.iter_rows(min_row=2, max_row=styled_sheet.max_row, min_col=1,
                                              max_col=styled_sheet.max_column):
                row_data = [cell.value for cell in row]
                data.append(row_data)

            # Define the columns for the DataFrame (assuming you have column names)
            column_names = [cell.value for cell in styled_sheet.iter_rows(min_row=1, max_row=1, min_col=1,
                                                                          max_col=styled_sheet.max_column).__next__()]

            # Create the DataFrame
            styled_df = pd.DataFrame(data, columns=column_names)
            # Perform inner join
            inner_join_result = pd.merge(sheet1, sheet2, how='inner')
            inner_join_result.to_excel(writer, sheet_name=f'i{sheet_name}', index=False)

            new_sync = pd.concat([inner_join_result, styled_df], ignore_index=True)

            new_sync_data[sheet_name] = new_sync
            # Create a separate Excel writer for the new_sync DataFrame



            print(f"Joins completed for sheet: {sheet_name}")
        except Exception as e:
            print(f"Error processing sheet '{sheet_name}': {e}")

print("All joins completed.")
#Change the file path to the sync path file
sync_path = fr"C:\Users\eg724520\Valmont Industries, Inc\Global IT AppSOD & DP - Documents\Team Tools\PyhtonCode - SAP Workstream DDF Data Refresh\Level 3 Tracking Data\Output\sync_results_Level3TrackingData_{date_time}.xlsx"
with pd.ExcelWriter(sync_path, engine='openpyxl') as final_writer:
    for sheet in new_sync_data.keys():
        sheet_df = new_sync_data[sheet]
        del sheet_df['_merge']
        sheet_df.to_excel(final_writer, sheet_name=f'{sheet}', index = False)



print("All new_sync DataFrames saved in a single workbook with multiple sheets.")